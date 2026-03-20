import argparse
import getpass
import re
from typing import List, Dict, Optional,Any
from netmiko import ConnectHandler
from pprint import pprint
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
SHOW_INT_DESC_CMD = "show interfaces description"


import re
from typing import List, Dict

def parse_show_int_desc(output: str) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = [] ## this creates a list of dictionaries that is empty 

    for line in output.splitlines(): ## splitlines differentiates new line
        line = line.rstrip()        ##rstrip() removes whitespace 
        if not line:
            continue

        low = line.lower()          ##lowercases the line 
        if low.startswith("interface") and "protocol" in low: 
            continue                                    ### these 2 skip to the real output; no int status protocol description line 
        if line.startswith(("R1(", "R2(", "R3(")):
            continue

        # Split into at most 4 columns: intf, status, proto, desc
        parts = re.split(r"\s{2,}", line, maxsplit=3)           ###here I split every line in columns; if 2 spaces are encountered ->> next column
        if len(parts) < 3:                                      ### maxsplit 3 ->> ensures I don t lose description 
            continue

        intf = parts[0].strip()                         ### interface name
        status = parts[1].strip()                       ### interface status
        proto = parts[2].strip()                        ### interface protocol
        desc = parts[3].strip() if len(parts) == 4 else ""  ### description ; makes description "" if len has only 3 argumetns(no description )
        desc = desc.replace("\u00A0", " ").lstrip(" -\t")
        rows.append({
            "interface": intf,
            "status": status,                       ### appens dictionaries to list 
            "description": desc,   
            # "protocol": proto,
        })

    return rows

def find_physical_interfaces(output:str) -> List[str]:
    phys_interfaces: List[str] = []
    for line in output.splitlines():
        line= line.rstrip()
        if not line:
            continue

        low = line.lower()          ##lowercases the line 
        if low.startswith("interface") and "protocol" in low: 
            continue                                    ### these 2 skip to the real output; no int status protocol description line 
        if line.startswith(("R1(", "R2(", "R3(")):
            continue

        parts = re.split(r"\s{2,}", line, maxsplit=3)
        intf = parts[0].strip()

        if "." in intf or low.startswith("lo") or low.startswith("nv"):
             continue
        phys_interfaces.append(intf)
    return phys_interfaces

def get_mac_interfaces(conn, interfaces: List[str]) -> List[dict]:
    mac_re = re.compile(r"\baddress is\s+(?P<mac>[0-9a-fA-F.]{14})\b")

    result = []
    for intf in interfaces:
        out = conn.send_command(f"show interfaces {intf}")
        m = mac_re.search(out)
        mac = m.group("mac").lower() if m else None
        result.append({"interface": intf, "mac": mac})

    return result
    
def merger(description_rows,mac_rows):
    mac_by_interface={d["interface"]:d["mac"] for d in mac_rows}
    for row in description_rows:
        intf=row["interface"]
        parent=intf.split(".",1)[0]
        row["mac"]= mac_by_interface.get(parent)
    return description_rows

def collect_router_report(conn, host_label):
    raw_output = conn.send_command(SHOW_INT_DESC_CMD)
    inventory = conn.send_command("show inventory")

    rows = parse_show_int_desc(raw_output)
    physical = find_physical_interfaces(raw_output)
    mac_rows = get_mac_interfaces(conn, physical)

    merged = merger(rows, mac_rows)

    return merged, inventory
    
def ssh_connect(
    host: str,
    username: str,
    password: str,
    secret: Optional[str] = None,
    device_type: str = "cisco_ios",):

    device = {
        "device_type": device_type,
        "host": host,
        "username": username,
        "password": password, }
    if secret:                      ### only used if we need a secret to get past enable.
        device["secret"] = secret

    conn = ConnectHandler(**device)
    if secret:
        conn.enable()

    return conn

# def write_xlsx_multi_sheets(router_sections: List[Dict[str, Any]], filename: str = "interfaces.xlsx") -> None:
    """
    One Excel file, multiple sheets.
    Each sheet: interface table + 3 blank rows + raw 'show inventory' output.
    """
    wb = Workbook()
    # remove default sheet
    default_ws = wb.active
    wb.remove(default_ws)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    inv_font = Font(name="Consolas")
    inv_align = Alignment(wrap_text=False)

    for idx, section in enumerate(router_sections, start=1):
        sheet_name = str(section["sheet"])[:31]  # Excel limit
        rows: List[Dict[str, Any]] = section.get("rows", [])
        inventory_output: str = section.get("inventory", "")

        ws = wb.create_sheet(title=sheet_name)

        # ---- Interface table ----
        if rows:
            # Ensure comments exists
            for r in rows:
                r.setdefault("comments", "")

            # Force comments last
            headers = list(rows[0].keys())
            if "comments" in headers and headers[-1] != "comments":
                headers = [h for h in headers if h != "comments"] + ["comments"]
            elif "comments" not in headers:
                headers.append("comments")

            # Write headers
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align

            # Write data
            for r_i, row in enumerate(rows, start=2):
                for c_i, h in enumerate(headers, start=1):
                    v = row.get(h, "")
                    ws.cell(row=r_i, column=c_i, value="" if v is None else v)

            end_row = 1 + len(rows)
            end_col = len(headers)
            ref = f"A1:{get_column_letter(end_col)}{end_row}"

            # unique table name per sheet
            tbl = Table(displayName=f"InterfacesTable{idx}", ref=ref)
            tbl.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium9",
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(tbl)

            ws.freeze_panes = "A2"

            # simple width autosize
            for c in range(1, end_col + 1):
                col_letter = get_column_letter(c)
                max_len = max(len(str(ws.cell(row=r, column=c).value or "")) for r in range(1, end_row + 1))
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 60)

            start_inv = end_row + 4  # 3 blank rows
        else:
            ws.cell(row=1, column=1, value="(no interface data)")
            start_inv = 5

        # ---- Inventory section ----
        ws.cell(row=start_inv, column=1, value="show inventory").font = Font(bold=True, size=14)

        for r, line in enumerate((inventory_output or "").splitlines(), start=start_inv + 1):
            cell = ws.cell(row=r, column=1, value=line.rstrip("\n"))
            cell.font = inv_font
            cell.alignment = inv_align

        # wider column A for inventory text
        ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 12, 80)

    wb.save(filename)

def main():
    # Only keep CLI args for connection details
    parser = argparse.ArgumentParser(
        description="Learning version: run a hardcoded command and print parsed output."
    )
    parser.add_argument("--hosts",nargs="+", required=True, help="Router IP/hostname")
    parser.add_argument("--username", required=True, help="SSH username")
    parser.add_argument("--password", help="SSH password (if omitted, you will be prompted)")
    parser.add_argument("--secret", help="Enable secret (optional)")

    args = parser.parse_args()
    router_sections: List[Dict[str, Any]] = []


    # Prompt for password if not given
    password = args.password or getpass.getpass("SSH Password: ")

    for host in args.hosts:
        conn = ssh_connect(
            host=host,
            username=args.username,
            password=password,
            secret=args.secret,
            device_type="cisco_ios",
        )
        try:
            result = collect_router_report(conn, host_label=host)
            merged = result[0]
            inventory = result[1]
            router_sections.append(
                {
                    "sheet": f"R_{host}",
                    "rows": merged,
                    "inventory": inventory,
                }
            )
        finally:
            conn.disconnect()

    write_xlsx_multi_sheets(router_sections, filename="sankyy.xlsx")
    print("Wrote interfaces.xlsx")


if __name__ == "__main__":
    main()