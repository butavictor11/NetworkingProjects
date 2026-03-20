import argparse
import getpass
import csv
import os
import re
from typing import List, Dict, Any, Iterable, Optional
from typing import List, Dict, Optional,Any
from netmiko import ConnectHandler
from pprint import pprint
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
def safe_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]+', "_", name).strip()

def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)

def write_dicts_csv(
    filename: str,
    rows: List[Dict[str, Any]],
    fieldnames: List[str],
    *,
    append: bool = False,
) -> str:
    """
    Writes rows (list of dicts) to CSV with explicit column order (fieldnames).
    If append=True, appends and only writes header if file does not exist / is empty.
    """
    mode = "a" if append else "w"
    file_exists = os.path.exists(filename)

    with open(filename, mode, newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        write_header = True

        if append and file_exists:
            # if file has content, don't rewrite header
            f.seek(0, os.SEEK_END)
            if f.tell() > 0:
                write_header = False

        if write_header:
            writer.writeheader()

        for r in rows:
            # fill missing columns with blank
            out = {k: r.get(k, "") for k in fieldnames}
            writer.writerow(out)

    return filename
def write_host_csvs(
    host: str,
    merged_info: List[Dict[str, Any]],
    vlan_table: List[Dict[str, Any]],
    inventory_output: str,
    out_dir: str = "csv_out",
) -> tuple[str, str]:
    ensure_dir(out_dir)
    host_safe = safe_filename(host)

    interfaces_csv = os.path.join(out_dir, f"{host_safe}_interfaces.csv")
    vlans_csv = os.path.join(out_dir, f"{host_safe}_vlans.csv")

    write_dicts_csv(
        interfaces_csv,
        merged_info,
        fieldnames=["interface", "status", "description", "mac", "comments"],
        append=False,
    )

    append_inventory_section(interfaces_csv, inventory_output, blank_lines=3)

    write_dicts_csv(
        vlans_csv,
        vlan_table,
        fieldnames=["vlan", "vlan_name", "mac", "aph", "comments"],
        append=False,
    )

    return interfaces_csv, vlans_csv
def append_inventory_section(csv_file: str, inventory_output: str, blank_lines: int = 3) -> None:
    with open(csv_file, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        for _ in range(blank_lines):
            w.writerow([])
        w.writerow(["show inventory"])
        for line in (inventory_output or "").splitlines():
            w.writerow([line])
def safe_sheet_name(name: str) -> str:
    bad = set('[]:*?/\\')
    out = "".join("_" if c in bad else c for c in name).strip()
    return (out[:31] if out else "Sheet")

def unique_sheet_name(wb: Workbook, desired: str) -> str:
    base = safe_sheet_name(desired)
    name = base
    i = 1
    existing = {ws.title for ws in wb.worksheets}
    while name in existing:
        suffix = f"_{i}"
        name = (base[:31 - len(suffix)] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        i += 1
    return name

def _style_header(ws, row_idx: int, end_col: int):
    for c in range(1, end_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN

def _autosize_columns(ws, max_width: int = 60):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def _write_table_only(ws, rows: List[Dict[str, Any]], columns: List[str], table_name: str) -> int:
    """
    Writes an Excel Table starting at row 1.
    Returns the last row number of the table (end_row).
    """
    # ensure comments column exists
    if "comments" not in columns:
        columns = columns + ["comments"]

    # header
    ws.append(columns)
    _style_header(ws, 1, len(columns))

    # data
    for r in rows:
        # normalize missing/None
        out_row = []
        for c in columns:
            v = r.get(c, "")
            out_row.append("" if v is None else v)
        ws.append(out_row)

    end_row = ws.max_row
    end_col = len(columns)

    # align cells
    for row in ws.iter_rows(min_row=2, max_row=end_row, min_col=1, max_col=end_col):
        for cell in row:
            cell.alignment = CELL_ALIGN

    # add Excel table object (only if there are data rows)
    if end_row >= 2:
        ref = f"A1:{get_column_letter(end_col)}{end_row}"
        tab = Table(displayName=table_name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tab)
        ws.freeze_panes = "A2"

    _autosize_columns(ws)
    return end_row

def _append_inventory_block(ws, start_row: int, inventory_output: str, blank_lines: int = 3):
    """
    Appends:
      <blank lines>
      show inventory
      <inventory lines in col A>
    starting from start_row+1 basically.
    """
    r = start_row + 1

    # blank rows
    for _ in range(blank_lines):
        ws.cell(row=r, column=1, value="")
        r += 1

    # title
    title_cell = ws.cell(row=r, column=1, value="show inventory")
    title_cell.font = Font(bold=True, size=14)
    r += 1

    inv_font = Font(name="Consolas")
    inv_align = Alignment(wrap_text=False)

    for line in (inventory_output or "").splitlines():
        cell = ws.cell(row=r, column=1, value=line.rstrip("\n"))
        cell.font = inv_font
        cell.alignment = inv_align
        r += 1

    # make column A wide enough for inventory
    ws.column_dimensions["A"].width = max(ws.column_dimensions["A"].width or 12, 80)

def create_report_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def add_host_sheets(
    wb: Workbook,
    host: str,
    merged_info: List[Dict[str, Any]],
    vlan_table: List[Dict[str, Any]],
    inventory_output: str,
    idx: int,
):
    # ---- Interfaces sheet ----
    ws_int = wb.create_sheet(unique_sheet_name(wb, f"{host}_interfaces"))
    end_row = _write_table_only(
        ws_int,
        rows=merged_info,
        columns=["interface", "status", "description", "mac", "comments"],
        table_name=f"IntTbl_{idx}",
    )
    _append_inventory_block(ws_int, start_row=end_row, inventory_output=inventory_output, blank_lines=3)

    # ---- VLAN sheet ----
    ws_vlan = wb.create_sheet(unique_sheet_name(wb, f"{host}_vlans"))
    _write_table_only(
        ws_vlan,
        rows=vlan_table,
        columns=["vlan", "vlan_name", "mac", "aph", "comments"],
        table_name=f"VlanTbl_{idx}",
    )

def save_report(wb: Workbook, filename: str = "network_report.xlsx") -> str:
    wb.save(filename)
    return filename
def ssh_connect(
    host: str,
    username: str,
    password: str,
    secret: Optional[str] = None,
    device_type: str = "cisco_ios",):

    device = {
        "device_type": device_type,
        "host": host,
        "username": username,
        "password": password, }
    if secret:                      ### only used if we need a secret to get past enable.
        device["secret"] = secret

    conn = ConnectHandler(**device)
    if secret:
        conn.enable()

    return conn
def parse_int_status(int_status):
    output:List[dict[str,str]]=[]
    physical_interfaces:List[str]=[]
    for line in int_status.splitlines():
        line=line.rstrip()
        if not line:
            continue
        low_line=line.lower()
        if low_line.startswith("interface"):
            continue
        if low_line.startswith(("vl", "vlan")):
            continue
        parts=re.split(r"\s+",line,maxsplit=3)
        interface=parts[0].strip()
        if "." not in interface and "Vl" not in interface:
            physical_interfaces.append(interface)
        status=parts[1].strip()
        proto=parts[2].strip()
        description=parts[3].strip() if len(parts) > 3 else ""
        description= description.replace("\u00A0", " ").lstrip(" -\t")
        if status.lower() == "admin" and proto.lower() == "down":
            status = "admin down"

            # desc currently looks like: "down ...." or just "down" or already pure description
            rest = description.strip()
            if rest:
                first, *tail = rest.split(None, 1)  # split once on any whitespace
                if first.lower() in ("up", "down"):
                    description = tail[0] if tail else ""  # remove leading proto
                else:
                    description = rest
            else:
                description = ""
        output.append({
            "interface":interface,
            "status":status,
            "description":description
        })
    return output,physical_interfaces
    
def parse_vlan_info(vlan_output):
    rows_table:List[Dict[str,str]]=[]
    vlan_list:List["str"]=[]
    for line in vlan_output.splitlines():
        line=line.rstrip()
        if not line:
            continue
        low_line=line.lower()
        if low_line.startswith("vlan") and "name" in low_line:
            continue
        if low_line.startswith("---"):
            continue
        if low_line.startswith(("1 ", "1002", "1003", "1004","1005")):
            continue
        if low_line.startswith("vlan type"):
            break
            
        first=re.split(r"\s+",line,maxsplit=2)
        vlan_id=first[0].strip()
        rest=first[1]
        name = re.split(r"\s{2,}", rest, maxsplit=1)[0].strip()
        rows_table.append({
        "vlan" : vlan_id,
        "vlan_name" : name,
        })
        vlan_list.append(vlan_id)

    return rows_table,vlan_list
def collect_info(conn):
    vlan_output=conn.send_command("sh vlan")
    int_status=conn.send_command("sh int description")
    inventory=conn.send_command("sh inventory ")
    return vlan_output,int_status,inventory


MAC_ROW_RE = re.compile(
    r"^\s*(?P<vlan>\d+)\s+(?P<mac>[0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4})\s+\S+\s+(?P<port>\S+)\s*$"
)

def inspect_vlan_output(vlan_list, conn, aph_interface:str):
    mac_aph: List[Dict[str,str]] = []
    aph_interface_standardized=aph_interface.strip().lower()
    for vlan in vlan_list:
        output=conn.send_command(f"show mac address-table dynamic vlan {vlan}")
        saw_mac=False
        saw_aph=False

        for line in output.splitlines():
            m=MAC_ROW_RE.match(line)

            if not m:
                continue
            saw_mac= True
            if m.group("port").strip().lower() == aph_interface_standardized:
                saw_aph= True
        if not saw_mac:
            mac_aph.append({"vlan":vlan, "mac": "no","aph":"no"})
        else:
            mac_aph.append({"vlan":vlan , "mac" :"yes", "aph":"yes" if saw_aph else "no"})

    return mac_aph

def collect_mac(interfaces,conn):
    mac_list:List[Dict[str,str]]=[]
    mac_re = re.compile(r"\baddress is\s+(?P<mac>[0-9a-fA-F.]{14})\b")
    for int in interfaces:
        output=conn.send_command(f"sh interface {int}")
        m=mac_re.search(output)
        mac=m.group("mac").lower() if m else None
        mac_list.append({
            "interface":int,
            "mac":mac,
        })
    return mac_list

def determine_shutdown(interfaces,conn):
    shut_interface:List[Dict[str,str]]=[]
    for int in interfaces:
        output=conn.send_command(f"sh run interface {int}")
        shut_interface.append({
                "interface":int,
                "comments":"shutdown" if "shutdown" in output.lower() else ""
            })
    return shut_interface
def merger1(description_rows,mac_rows):
    mac_by_interface={d["interface"]:d["mac"] for d in mac_rows}
    for row in description_rows:
        intf=row["interface"]
        parent=intf.split(".",1)[0]
        row["mac"]= mac_by_interface.get(parent)
    return description_rows
def merger_comm(description_rows,shut_rows):
    shut_by_interface={d["interface"]:d["comments"] for d in shut_rows}
    for row in description_rows:
        intf=row["interface"]
        parent=intf.split(".",1)[0]
        row["comments"]= shut_by_interface.get(parent)
    return description_rows
def merger2(description_rows, mac_rows):
    mac_by_vlan = {str(d["vlan"]): d for d in mac_rows}

    merged = []
    for row in description_rows:
        vlan = str(row.get("vlan") or row.get("vlan_id"))
        m = mac_by_vlan.get(vlan, {})

        merged.append({
            "vlan": vlan,
            "vlan_name": row.get("vlan_name", ""),
            "mac": m.get("mac", ""),
            "aph": m.get("aph", ""),
        })

    return merged
def main():
    # Only keep CLI args for connection details
    parser = argparse.ArgumentParser(
        description="Learning version: run a hardcoded command and print parsed output."
    )
    parser.add_argument("--hosts",nargs="+", required=True, help="Router IP/hostname")
    parser.add_argument("--username", required=True, help="SSH username")
    parser.add_argument("--password", help="SSH password (if omitted, you will be prompted)")
    parser.add_argument("--secret", help="Enable secret (optional)")
    parser.add_argument("--waninterface",help="interface to aph router")

    args = parser.parse_args()
    password = args.password or getpass.getpass("SSH Password: ")
    wb = create_report_workbook()
    idx = 0  # used for unique table names

    for host in args.hosts:
        idx += 1
        conn = ssh_connect(
            host=host,
            username=args.username,
            password=password,
            secret=args.secret,
            device_type="cisco_ios",
        )
        try:
            vlan_output, int_status, inventory = collect_info(conn)

            parsed_vlan, vlan_list = parse_vlan_info(vlan_output)

            interface_status, physical_interfaces = parse_int_status(int_status)
            mac_dictionary = collect_mac(physical_interfaces, conn)
            merged_info = merger1(interface_status, mac_dictionary)
            shutdown=determine_shutdown(physical_interfaces,conn)
            merged_info=merger_comm(merged_info,shutdown)
            mac_use = inspect_vlan_output(vlan_list, conn, args.waninterface)
            vlan_table = merger2(parsed_vlan, mac_use)
            # print(merged_info2)
            for r in merged_info:
                r["comments"] = r.get("comments") or ""
                r["mac"] = r.get("mac") or ""
            for r in vlan_table:
                r["comments"] = r.get("comments") or ""

            add_host_sheets(wb, host, merged_info, vlan_table, inventory, idx)
            int_csv, vlan_csv = write_host_csvs(host, merged_info, vlan_table, inventory, out_dir="CSV")
            print(f"[{host}] wrote {int_csv}")
            print(f"[{host}] wrote {vlan_csv}")

            # create 2 sheets for this host and write the tables
            # add_host_sheets(wb, host, merged_info, vlan_table, idx)

        finally:
            conn.disconnect()

    out_xlsx = save_report(wb, "network_report.xlsx")
    print(f"Wrote Excel: {out_xlsx}")

if __name__ == "__main__":
    main()